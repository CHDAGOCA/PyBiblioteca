import datetime
import re
import csv
import os
import openpyxl
import sqlite3
import sys
import os.path
from sqlite3 import Error

def registro():
    print("-" * 40)
    print("Registro de libro")
    print("-" * 40)
    while True:
            while True:
                    while True: 
                        titulo=input("Ingresa el titulo del libro a registrar: \n")
                        if titulo.strip() == '': 
                            print("El titulo es un campo obligatorio ")
                        else: 
                            break
                            
                    while True:
                        ConsultaLibro_TAG(0,1,0)
                        autor=input(f"Ingrese el autor de: {titulo} \n")
                        if autor.strip() == '':
                            print("El autor del libro es un campo obligatorio")
                        elif ChecarAut(autor)==False:
                            print("\nAutor no registrado en base de datos.")
                            continue
                        else: 
                            break

                    while True:
                        ConsultaLibro_TAG(0,0,1)
                        genero=input(f"Ingrese el genero al que pertenece: \n")
                        if genero.strip() == '':
                            print("El genero del libro es un campo obligatorio")
                        elif ChecarGen(genero)==False:
                            print("\nGenero no registrado en base de datos.")
                            continue
                        else: 
                            break

                    while True:
                        publicacion=input(f"Ingrese el año de publicación del libro: {titulo} (YYYY)\n ")
                        if (not bool(re.match("^[0-9]{4}$", publicacion))):
                            print("\nEl año de publicación del libro solo pueden ser 4 caracteres númericos.")
                            continue
                        fecha_procesada = datetime.datetime.strptime(publicacion, "%Y").date() 
                        fecha_actual = datetime.date.today()
                        #print(fecha_procesada,type(fecha_procesada))
                        if fecha_procesada > fecha_actual:
                            print("Esta fecha no es valida")
                        else:
                            break

                    while True:
                        fechadq_captura=input("Ingrese la fecha en la que se adquirio el libro [dd/mm/aaaa] \n")
                        fecha_adquisicion=fechadq_captura[6:10]+"/"+fechadq_captura[3:5]+"/"+fechadq_captura[:2]
                        if (not bool(re.match("^([0-9]{4}[/]?((0[13-9]|1[012])[/]?(0[1-9]|[12][0-9]|30)|(0[13578]|1[02])[/]?31|02[/]?(0[1-9]|1[0-9]|2[0-8]))|([0-9]{2}(([2468][048]|[02468][48])|[13579][26])|([13579][26]|[02468][048])00)[/]?02[/]?29)$",fecha_adquisicion))):
                            print("\nLa fecha sigue los formatos aaaa/mm/dd y solo acepta dias posibles.")
                            continue
                        fecha2_procesada= datetime.datetime.strptime(fecha_adquisicion, "%Y/%m/%d").date() 
                        if fecha2_procesada < fecha_procesada :
                            print("La fecha de adquisicion debe ser despues de la fecha de publicacion, ingrese una fecha valida ")
                        else:
                            break

                    while True:
                        isbn=input(str(f"Ingresa la clave de ISBN del libro \n"))
                        if len(isbn) == 13 and (bool(re.match("^[0-9]{13}$", isbn))):
                            break
                        else:
                            print("El ISBN debe tener 13 caracteres numericos, vuelva a ingresarlos ")

                    while True:
                            DatosCorrect=True
                            print("*" * 40)
                            print(f"Titulo: {titulo} \nAutor: {autor} \nGenero: {genero} \nFecha de publicacion: {publicacion} \nFecha de adquisición: {fecha_adquisicion} \nISBN: {isbn} \n") 
                            print("*" * 40)
                            confirmacion=input("¿Son correctos los datos ingresados? (SI/NO)\n ").upper()
                            print("*" * 40)
                            if confirmacion=="SI":
                                print("*" * 40)
                                GuardarLibros(titulo,Obt_CL_Aut(autor),Obt_CL_Gen(genero),int(publicacion),isbn,int(fecha_adquisicion[8:10]),int(fecha_adquisicion[5:7]),int(fecha_adquisicion[:4]))
                                print(f"Se registro el libro\n")
                                print("*" * 40)
                                break
                            elif confirmacion=="NO": 
                                DatosCorrect=False
                                print("Vuelva a ingresar los datos")
                                break
                            else:
                                print("Introduce una de las opciones (SI/NO)\n ")
                    if DatosCorrect==False:
                        continue   

                    break
            seleccion=False
            while True:
                nuevo_registro=input("¿Deseas realizar un nuevo registro? (SI/NO)\n ").upper()
                if nuevo_registro=="SI":
                    seleccion=True
                    break 
                elif nuevo_registro=="NO":
                    print("*" * 40)
                    print("Sus registros han quedado guardados")
                    break
                else:
                    print("Favor de seleccionar una opcion valida.")
                    continue
            if seleccion==True:
                continue
            else:
                break

def consultas():
    while True:
        try:
            print("Consultas y Reportes")
            print("[1] - Consulta por titulo \n[2] - Reportes \n[3] - Volver al menú Principal \n")
            sub_menu=input("¿Que accion deseas realizar?\n ")
        
            if sub_menu == "3":
                break
            if sub_menu == "1":
                while True:
                    print("Consulta de Título")
                    print("[1] - Por Título \n[2] - Por ISBN \n[3] - Volver al menú de consultas y reportes ")
                    consulta = input("¿Qué acción deseas realizar?\n ")
                    if consulta == "3":
                        break
                    if consulta == "1":#Por Titulo
                        try:
                            with sqlite3.connect("Biblioteca.db") as conn:
                                mi_cursor = conn.cursor()

                                mi_cursor.execute("SELECT titulo FROM Libros")
                                registros = mi_cursor.fetchall()

                                if registros:
                                    print("**********Lista de Títulos*********")
                                    for titulo in registros:
                                        print(titulo[0])
                                    print("*" * 35)

                                buscar_titulo = input("¿Qué título quieres consultar? \n").upper()
                                valores = {"titulo": buscar_titulo}

                                datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                        FROM Libros \
                                        JOIN autores ON Libros.autor = autores.clave \
                                        JOIN generos ON Libros.genero = generos.clave \
                                        WHERE Libros.titulo = :titulo"

                                mi_cursor.execute(datos, valores)
                                registros2 = mi_cursor.fetchall()

                                if registros2:
                                    print("**********Resultados de la búsqueda*********")
                                    for fila in registros2:
                                        print("Clave: ", fila[0])
                                        print("Título: ", fila[1])
                                        print("Autor: ", fila[2], fila[3])
                                        print("Género: ", fila[4])
                                        print("Año de publicacion: ", fila[5])
                                        print("ISBN: ", fila[6])
                                        print("Fecha en la que se adquirio: ", fila[7])
                                else:
                                    print("No se encontró el libro")

                        except Error as e:
                            print(e)
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()


                    if consulta == "2":#Por ISBN
                        try:
                            with sqlite3.connect("Biblioteca.db") as conn:
                                mi_cursor = conn.cursor()

                                mi_cursor.execute("SELECT ISBN FROM Libros")
                                registros = mi_cursor.fetchall()

                                if registros:
                                    print("**********Lista de ISBN*********")
                                    for isbn in registros:
                                        print(isbn[0])
                                    print("*" * 35)

                                while True:
                                    buscar_isbn=input(str(f"Ingresa la clave de ISBN del libro\n "))
                                    if len(buscar_isbn) == 13 and (bool(re.match("^[0-9]{13}$", buscar_isbn))):
                                        break
                                    else:
                                        print("El ISBN debe tener 13 caracteres numericos, vuelva a ingresarlos ")
                                valores2 = {"isbn": buscar_isbn}

                                datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                            FROM Libros \
                                            JOIN autores ON Libros.autor = autores.clave \
                                            JOIN generos ON Libros.genero = generos.clave \
                                            WHERE Libros.ISBN = :isbn"

                                mi_cursor.execute(datos, valores2)
                                registros2 = mi_cursor.fetchall()

                                if registros2:
                                    print("**********Resultados de la búsqueda*********")
                                    for fila in registros2:
                                        print("Clave: ", fila[0])
                                        print("Título: ", fila[1])
                                        print("Autor: ", fila[2], fila[3])
                                        print("Género: ", fila[4])
                                        print("Año de publicacion: ", fila[5])
                                        print("ISBN: ", fila[6])
                                        print("Fecha en la que se adquirio: ", fila[7])
                                else:
                                    print("No se encontró el libro")

                        except Error as e:
                            print(e)
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                            
            if sub_menu == "2":
                while True:
                    print("REPORTES")
                    print("[1] - Catalogo completo \n[2] - Reporte por Autor \n[3] - Reporte por Genero \n[4] - Reporte por Año de publicación\n[5] - Volver al menu de reportes")
                    opcion = input("¿Qué acción deseas realizar? \n ")
                    if opcion == "1":
                            try:
                                with sqlite3.connect("Biblioteca.db") as conn:
                                    mi_cursor = conn.cursor()
                                    registros = mi_cursor.fetchall()

                                    datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                            FROM Libros \
                                            JOIN autores ON Libros.autor = autores.clave \
                                            JOIN generos ON Libros.genero = generos.clave"

                                    mi_cursor.execute(datos)
                                    registros2 = mi_cursor.fetchall()


                                    if registros2:
                                        print("**********Resultados de la búsqueda*********")
                                        print("Titulo  -  Nombre del autor  -  Apellido del autor  -  Genero  -  Año de publicacion  -   ISBN   -   Fecha de Adquisicion")
                                        for fila in registros2:
                                            print(f"{fila[1]} || {fila[2]} {fila[3]} || {fila[4]} || {fila[5]} || {fila[6]} || {fila[7]} ")
                                    else:
                                        print("No se encontró el libro")

                                    while True:
                                        print("Formas de exportación\n[1] - Exportar a CSV \n[2] - Exportar a msExcel \n[3] - No exportar Reporte")
                                        exportarP = input("¿Que desea hacer? \n")
                                        if exportarP=="1":
                                            GenArch_CatComp_CSV()
                                            continue
                                        elif exportarP=="2":
                                            GenArch_CatComp_Excel()
                                            continue
                                        elif exportarP=="3":
                                            break
                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()

                    if opcion == "2":
                            try:
                                with sqlite3.connect("Biblioteca.db") as conn:
                                    mi_cursor = conn.cursor()

                                    mi_cursor.execute("SELECT clave, AutNombre, AutApellidos FROM autores")
                                    registros = mi_cursor.fetchall()

                                    if registros:
                                        print("**********Lista de Autores*********")
                                        for clave,autnombre,autapellidos in registros:
                                            print("Clave/Nombre/Apelido")
                                            print(clave,autnombre,autapellidos)
                                        print("*" * 35)
                                    while True:
                                        buscar_autor=input(f"Ingrese el nombre completo del autor: \n")
                                        if buscar_autor.strip() == '':
                                            print("El autor del libro es un campo obligatorio")
                                        elif ChecarAut(buscar_autor)==False:
                                            print("\nAutor no registrado en base de datos.")
                                            continue
                                        else: 
                                            break
                                    valores2 = {"autor": buscar_autor.upper()}

                                    datos = "SELECT Libros.clave, Libros.titulo, Libros.añopublicacion, autores.AutNombre, autores.AutApellidos \
                                                FROM Libros \
                                                JOIN autores ON Libros.autor = autores.clave \
                                                JOIN generos ON Libros.genero = generos.clave \
                                                WHERE (autores.AutNombre||' '||autores.AutApellidos) = :autor"

                                    mi_cursor.execute(datos, valores2)
                                    registros3 = mi_cursor.fetchall()

                                    if registros3:
                                        print("**********Resultados de la búsqueda*********")
                                        print("Titulo/Fecha de publicacion")
                                        for fila in registros3:
                                            print(f"{fila[1]}, | {fila[2]}")
                                        print("*" * 35)
                                    else:
                                        print("No se encontró el libro")
                                    
                                    while True:
                                        print("Formas de exportación\n[1] - Exportar a CSV \n[2] - Exportar a msExcel \n[3] - No exportar Reporte")
                                        exportarP = input("¿Que desea hacer? \n")
                                        if exportarP=="1":
                                            GenArch_CatAut_CSV(buscar_autor)
                                        elif exportarP=="2":
                                            GenArch_CatAut_Excel(buscar_autor)
                                        elif exportarP=="3":
                                            break

                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                    if opcion == "3":
                            try:
                                with sqlite3.connect("Biblioteca.db") as conn:
                                    mi_cursor = conn.cursor()

                                    mi_cursor.execute("SELECT clave, GenNombre FROM generos")
                                    registros = mi_cursor.fetchall()

                                    if registros:
                                        print("**********Lista de Generos*********")
                                        print("Clave/Nombre del genero")
                                        for clave,GenNombre in registros:
                                            print(clave,GenNombre)
                                        print("*" * 35)

                                    while True:
                                        buscar_autor=input(f"Ingrese el genero al que pertenece: \n")
                                        if buscar_autor.strip() == '':
                                            print("El genero del libro es un campo obligatorio")
                                        elif ChecarGen(buscar_autor)==False:
                                            print("\nGenero no registrado en base de datos.")
                                            continue
                                        else: 
                                            break
                                    valores2 = {"genero": buscar_autor.upper()}

                                    datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, Libros.añopublicacion \
                                                FROM Libros \
                                                JOIN autores ON Libros.autor = autores.clave \
                                                JOIN generos ON Libros.genero = generos.clave \
                                                WHERE generos.GenNombre = :genero"

                                    mi_cursor.execute(datos, valores2)
                                    registros3 = mi_cursor.fetchall()

                                    if registros3:
                                        print("**********Resultados de la búsqueda*********")
                                        print("Clave --- Titulo   ---   Nombre del autor   ---   Apellido del autor   ---   Fecha de publicacion")
                                        for fila in registros3:
                                            print(f"{fila[0]} {fila[1]} {fila[2]} {fila[3]} {fila[4]}")
                                    else:
                                        print("No se encontró el libro")

                                    while True:
                                        print("Formas de exportación\n[1] - Exportar a CSV \n[2] - Exportar a msExcel \n[3] - No exportar Reporte")
                                        exportarP = input("¿Que desea hacer? \n")
                                        if exportarP=="1":
                                            GenArch_CatGen_CSV(buscar_autor)
                                        elif exportarP=="2":
                                            GenArch_CatGen_Excel(buscar_autor)
                                        elif exportarP=="3":
                                            break


                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                    if opcion == "4":
                            try:
                                with sqlite3.connect("Biblioteca.db") as conn:
                                    mi_cursor = conn.cursor()
                                    while True:
                                        buscar_fecha=input(f"Ingrese el año de publicación del libro (YYYY): \n")
                                        if (not bool(re.match("^[0-9]{4}$", buscar_fecha))):
                                            print("\nEl año de publicación del libro solo pueden ser 4 caracteres númericos.")
                                            continue
                                        else:
                                            break
                                    fechaprocesada = datetime.datetime.strptime(buscar_fecha, "%Y").date()
                                    valores = {"fecha": fechaprocesada}

                                    datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                            FROM Libros \
                                            JOIN autores ON Libros.autor = autores.clave \
                                            JOIN generos ON Libros.genero = generos.clave \
                                            WHERE DATE(Libros.añopublicacion) = :fecha"

                                    mi_cursor.execute(datos, valores)
                                    registros2 = mi_cursor.fetchall()

                                    if registros2:
                                        print("**********Resultados de la búsqueda*********")
                                        print("Titulo   ---   Nombre del autor   ---   Apellido del autor   ---   Genero  ---  Año de publicacion   ---   ISBN")
                                        for fila in registros2:
                                            print(f"{fila[1]} {fila[2]} {fila[3]} {fila[4]} {fila[5]} {fila[6]}")
                                    else:
                                        print("No se encontró el libro")

                                    while True:
                                        print("Formas de exportación\n[1] - Exportar a CSV \n[2] - Exportar a msExcel \n[3] - No exportar Reporte")
                                        exportarP = input("¿Que desea hacer? \n")
                                        if exportarP=="1":
                                            GenArch_CatPubYear_CSV(fechaprocesada)
                                        elif exportarP=="2":
                                            GenArch_CatPubYear_Excel(fechaprocesada)
                                        elif exportarP=="3":
                                            break

                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                    if opcion == "5":
                        break
        except Error as e:
            print(e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

def CrearTablas():
    file_exists = os.path.exists('Biblioteca.db')
    if not(file_exists):
        try:
            with sqlite3.connect("Biblioteca.db") as conn:
                mi_cursor=conn.cursor()
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS generos (clave INTEGER PRIMARY KEY, GenNombre TEXT NOT NULL);")
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS autores (clave INTEGER PRIMARY KEY, AutNombre TEXT NOT NULL, AutApellidos TEXT NOT NULL);")
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS Libros (clave INTEGER PRIMARY KEY, titulo TEXT NOT NULL, autor INTEGER NOT NULL, \
                                  genero INTEGER NOT NULL, añopublicacion timestamp, ISBN TEXT NOT NULL, \
                                  fechaadq TIMESTAMP, FOREIGN KEY(autor) REFERENCES autores(clave), FOREIGN KEY(genero) REFERENCES genero(clave));")
                print("Tablas creada exitosamente")
        except Error as e:
                print(e)
        except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
                conn.close()
    else:
        print("Archivo db existente.")

#Funciones no utilizadas en esta version:
def ExportArchComplt_csv():
    listareport=list(registro_libro.items())
    nombrarch = "ReporteCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5]) for clave,datos in listareport])
    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchAutores_csv(autorsearch):
    listareport=list(registro_libro.items())
    nombrarch = "ReporteAutores" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    for clave,datos in listareport:
        if datos[1]==autorsearch:
            grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5])])
    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchGenero_csv(generosearch):
    listareport=list(registro_libro.items())
    nombrarch = "ReporteGenero" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    for clave,datos in listareport:
        if datos[2]==generosearch:
            grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5])])
    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchAñoPublic_csv(añosearch):
    listareport=list(registro_libro.items())
    nombrarch = "ReporteAñoPublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    
    for clave, datos in listareport:
        año = datos[3].split("/")[-1]  
        if año == añosearch:
            grabador1.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])])
    

    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchComplt_Excel():
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteCatalogoCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Catalogo completo"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        hoja.cell(row=i+2, column=2).value = clave
        hoja.cell(row=i+2, column=3).value = valor[0]
        hoja.cell(row=i+2, column=4).value = valor[1]
        hoja.cell(row=i+2, column=5).value = valor[2]
        hoja.cell(row=i+2, column=6).value = valor[3]
        hoja.cell(row=i+2, column=7).value = valor[4]
        hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def ExportArchAutores_Excel(autorsearch):
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteAutores" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Autores"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        if valor[1]==autorsearch:
            hoja.cell(row=i+2, column=2).value = clave
            hoja.cell(row=i+2, column=3).value = valor[0]
            hoja.cell(row=i+2, column=4).value = valor[1]
            hoja.cell(row=i+2, column=5).value = valor[2]
            hoja.cell(row=i+2, column=6).value = valor[3]
            hoja.cell(row=i+2, column=7).value = valor[4]
            hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def ExportArchAñoPublic_Excel(añosearch):
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteAñoPublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Año de Publicacion"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        if valor[3]==añosearch:
            hoja.cell(row=i+2, column=2).value = clave
            hoja.cell(row=i+2, column=3).value = valor[0]
            hoja.cell(row=i+2, column=4).value = valor[1]
            hoja.cell(row=i+2, column=5).value = valor[2]
            hoja.cell(row=i+2, column=6).value = valor[3]
            hoja.cell(row=i+2, column=7).value = valor[4]
            hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def GuardarLibros(titulo,autor,genero,añopub,isbn,fechadqdia,fechadqmes,fechañodq):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (titulo.upper(),autor,genero,datetime.datetime(añopub,1,1),isbn,datetime.datetime(fechañodq,fechadqmes,fechadqdia))
            mi_cursor.execute("INSERT INTO Libros (titulo,autor,genero,añopublicacion,ISBN,fechaadq) VALUES(?,?,?,?,?,?)", valores)
        print("Registros agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def GuardarAutores(nombre,apellidos):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (nombre.upper(),apellidos.upper())
            mi_cursor.execute("INSERT INTO autores (AutNombre,AutApellidos) VALUES(?,?)", valores)
        print("Autor agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def GuardarGeneros(genero):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (genero.upper(),)
            mi_cursor.execute("INSERT INTO generos (GenNombre) VALUES(?)", valores)
        print("Genero agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def registro_autores():
    while True:
        out=False
        autor=""
        apellidos=""
        while True:
            autor=input(f"Favor de ingresar el nombre del autor a registrar: \n")
            if autor.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",autor))):
                print("\nEl nombre del autor solo puede contener 100 caracteres como máximo entre letras y espacios.")
                continue
            else: 
                break
        while True:
            apellidos=input(f"Favor de ingresar los apellidos del autor a registrar: \n")
            if apellidos.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",apellidos))):
                print("\nLos apellidos del autor solo pueden contener 100 caracteres como máximo entre letras y espacios.")
                continue
            else: 
                break
        nomautor=autor+apellidos
        if ChecarAut(nomautor):
            print("Este Autor ya fue registrado previamente...")
            continue
        else:
            GuardarAutores(autor,apellidos)
        while True:
            print("Introduzca 1 para registrar otro autor\nIntroduzca 2 para salir de la seccion de registro de autores ")
            salida=input("Por favor elija una opción: \n")
            if salida=="1":
                out=False
                break
            elif salida=="2":
                out=True
                break
            else:
                print("Seleccion introducida no valida.")
        if out==True:
            break

def registro_generos():
    while True:
        out=False
        while True:
            genero=input(f"Favor de ingresar el genero a registrar: \n")
            if genero.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",genero))):
                print("\nEl nombre del genero solo puede contener 100 caracteres como máximo entre letras y espacios.")
                continue
            elif ChecarGen(genero):
                print("Este genero ya fue registrado previamente...")
                continue
            else: 
                break
            
        GuardarGeneros(genero)
        while True:
            print("Introduzca 1 para registrar otro autor\nIntroduzca 2 para salir de la seccion de registro de autores ")
            salida=input("Por favor elija una opción: \n")
            if salida=="1":
                out=False
                break
            elif salida=="2":
                out=True
                break
            else:
                print("Seleccion introducida no valida.")
        if out==True:
            break

def HayAutores():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def HayGeneros():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def HayLibros():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Libros ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def ConsultaLibro_TAG(titulo,autor,genero):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Libros ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if titulo==1:
                    print("titulos")
                    print("*" * 30)
                    for claves, titulos, autores, generos,añopub, isbn,fecha in registros:
                        print(f"{titulos:^16}")
        #Si no hay registros en la respuesta
            else:
                if titulo==1:
                    print("No hay libros registrados.")

            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if autor==1:
                    print("nombre\t\tapellidos")
                    print("*" * 30)
                    for claves, nombreaut,apellaut in registros:
                        print(f"{nombreaut:^16}\t{apellaut}")
        #Si no hay registros en la respuesta
            else:
                if autor==1:
                    print("No hay libros registrados.")
            
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if genero==1:
                    print("generos")
                    print("*" * 30)
                    for claves, generonom in registros:
                        print(f"{generonom:^16}")
        #Si no hay registros en la respuesta
            else:
                if genero==1:
                    print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def ChecarAut(autnom):
    found=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                for claves, nombreaut,apellaut in registros:
                    nombrecomp=nombreaut+" "+apellaut
                    if nombrecomp==autnom.upper():
                        found=True
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return found
    
def ChecarGen(genname):
    found=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                for claves, nombregen in registros:
                    if nombregen==genname.upper():
                        found=True
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return found

def Obt_CL_Aut(nomsearch):
    recover=""
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                for claves, nombreaut,apellaut in registros:
                    nombrecomp=nombreaut+" "+apellaut
                    if nombrecomp==nomsearch.upper():
                        recover=claves
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return recover
    
def Obt_CL_Gen(gensearch):
    recover=""
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                for claves, nombregen in registros:
                    if nombregen==gensearch.upper():
                        recover=claves
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return recover   

def GenArch_CatAut_CSV(search):
    nombrarch = "ReporteAutor" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"titulo": search.upper()}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave \
                    WHERE (autores.AutNombre||' '||autores.AutApellidos) = :titulo"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                for fila in registros2:
                    NomAutComp=fila[2]+' '+fila[3]
                    grabador1.writerows([(str(fila[0]),fila[1],NomAutComp,fila[4],fila[5],fila[7],fila[6])])
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        archivo4.close
        ruta = os.getcwd()
        print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)
        conn.close()

def GenArch_CatPubYear_CSV(search):
    nombrarch = ("ReporteAñodePublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv")
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"fecha": search}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                FROM Libros \
                                JOIN autores ON Libros.autor = autores.clave \
                                JOIN generos ON Libros.genero = generos.clave \
                                WHERE DATE(Libros.añopublicacion) = :fecha"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                for fila in registros2:
                    NomAutComp=fila[2]+' '+fila[3]
                    grabador1.writerows([(str(fila[0]),fila[1],NomAutComp,fila[4],fila[5],fila[7],fila[6])])
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        archivo4.close
        ruta = os.getcwd()
        print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)
        conn.close()

def GenArch_CatGen_CSV(search):
    nombrarch = "ReporteGenero" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"titulo": search.upper()}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave \
                    WHERE generos.GenNombre = :titulo"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                for fila in registros2:
                    NomAutComp=fila[2]+' '+fila[3]
                    grabador1.writerows([(str(fila[0]),fila[1],NomAutComp,fila[4],fila[5],fila[7],fila[6])])
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        archivo4.close
        ruta = os.getcwd()
        print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)
        conn.close()

def GenArch_CatComp_CSV():
    nombrarch = "ReporteCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave"

            mi_cursor.execute(datos)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                for fila in registros2:
                    NomAutComp=fila[2]+' '+fila[3]
                    grabador1.writerows([(str(fila[0]),fila[1],NomAutComp,fila[4],fila[5],fila[7],fila[6])])
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        archivo4.close
        ruta = os.getcwd()
        print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)
        conn.close()

def GenArch_CatAut_Excel(search):
    ruta = os.getcwd()
    archname = "ReporteCatalogoAutor" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte por Autor"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"titulo": search.upper()}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave \
                    WHERE (autores.AutNombre||' '||autores.AutApellidos) = :titulo"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                i=0
                for fila in registros2:
                    i=i+1
                    NomAutComp=fila[2]+' '+fila[3]
                    hoja.cell(row=i+1, column=2).value = str(fila[0])
                    hoja.cell(row=i+1, column=3).value = fila[1]
                    hoja.cell(row=i+1, column=4).value = NomAutComp
                    hoja.cell(row=i+1, column=5).value = fila[4]
                    hoja.cell(row=i+1, column=6).value = fila[5]
                    hoja.cell(row=i+1, column=7).value = fila[7]
                    hoja.cell(row=i+1, column=8).value = fila[6]
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        libro.save(archname)
        print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)
        conn.close()

def GenArch_CatGen_Excel(search):
    ruta = os.getcwd()
    archname = "ReporteCatalogoGenero" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte por Genero"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"titulo": search.upper()}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave \
                    WHERE generos.GenNombre = :titulo"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                i=0
                for fila in registros2:
                    i=i+1
                    NomAutComp=fila[2]+' '+fila[3]
                    hoja.cell(row=i+1, column=2).value = str(fila[0])
                    hoja.cell(row=i+1, column=3).value = fila[1]
                    hoja.cell(row=i+1, column=4).value = NomAutComp
                    hoja.cell(row=i+1, column=5).value = fila[4]
                    hoja.cell(row=i+1, column=6).value = fila[5]
                    hoja.cell(row=i+1, column=7).value = fila[7]
                    hoja.cell(row=i+1, column=8).value = fila[6]
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        libro.save(archname)
        print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)
        conn.close()

def GenArch_CatPubYear_Excel(search):
    ruta = os.getcwd()
    archname = "ReporteCatalogoAñodePublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte por Año de publicacion"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            valores = {"fecha": search}

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                                FROM Libros \
                                JOIN autores ON Libros.autor = autores.clave \
                                JOIN generos ON Libros.genero = generos.clave \
                                WHERE DATE(Libros.añopublicacion) = :fecha"

            mi_cursor.execute(datos, valores)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                i=0
                for fila in registros2:
                    i=i+1
                    NomAutComp=fila[2]+' '+fila[3]
                    hoja.cell(row=i+1, column=2).value = str(fila[0])
                    hoja.cell(row=i+1, column=3).value = fila[1]
                    hoja.cell(row=i+1, column=4).value = NomAutComp
                    hoja.cell(row=i+1, column=5).value = fila[4]
                    hoja.cell(row=i+1, column=6).value = fila[5]
                    hoja.cell(row=i+1, column=7).value = fila[7]
                    hoja.cell(row=i+1, column=8).value = fila[6]
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        libro.save(archname)
        print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)
        conn.close()

def GenArch_CatComp_Excel():
    ruta = os.getcwd()
    archname = "ReporteCatalogoCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Catalogo completo"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()

            datos = "SELECT Libros.clave, Libros.titulo, autores.AutNombre, autores.AutApellidos, generos.GenNombre, Libros.añopublicacion, Libros.ISBN, Libros.Fechaadq \
                    FROM Libros \
                    JOIN autores ON Libros.autor = autores.clave \
                    JOIN generos ON Libros.genero = generos.clave"

            mi_cursor.execute(datos)
            registros2 = mi_cursor.fetchall()
            
            if registros2:
                print("**********Resultados de la búsqueda*********")
                i=0
                for fila in registros2:
                    i=i+1
                    NomAutComp=fila[2]+' '+fila[3]
                    hoja.cell(row=i+1, column=2).value = str(fila[0])
                    hoja.cell(row=i+1, column=3).value = fila[1]
                    hoja.cell(row=i+1, column=4).value = NomAutComp
                    hoja.cell(row=i+1, column=5).value = fila[4]
                    hoja.cell(row=i+1, column=6).value = fila[5]
                    hoja.cell(row=i+1, column=7).value = fila[7]
                    hoja.cell(row=i+1, column=8).value = fila[6]
                    print("Clave: ", fila[0])
                    print("Título: ", fila[1])
                    print("Autor: ", fila[2], fila[3])
                    print("Género: ", fila[4])
                    print("Año de publicacion: ", fila[5])
                    print("ISBN: ", fila[6])
                    print("Fecha en la que se adquirio: ", fila[7])
            else:
                print("No se encontraron libros.")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        libro.save(archname)
        print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)
        conn.close()

CrearTablas()
while True:
    print("Bienvenido a la biblioteca universitaria")
    print("[1] - Registrar un nuevo ejemplar \n[2] - Consultas y reportes \
          \n[3] - Registrar un genero\n[4] - Registrar un autor\n[5] - Salir")
    menu_principal=input("¿Que accion deseas realizar?\n ")
    if menu_principal== "1":
        if HayAutores()==False and HayGeneros()==False:
            print("No hay autores, ni generos registrados por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        elif HayGeneros()==False:
            print("No hay generos registrados, por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        elif HayAutores()==False:
            print("No hay autores registrados, por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        else:
            registro()
    elif menu_principal=="2":
        if HayLibros()==False:
            print("No hay libros registrados, por lo que no hay libros para consultar o reportar.\nVolviendo a menu principal....")
        else:
            consultas()
    elif menu_principal=="5":
        print("Gracias por visitarnos, vuelva pronto")
        break
    elif menu_principal=="3":
        registro_generos()
    elif menu_principal=="4":
        registro_autores()
    else:
        print("La opcion ingresada no es correcta, elija de nuevo")
